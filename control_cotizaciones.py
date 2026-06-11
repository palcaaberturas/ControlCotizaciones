from __future__ import annotations

import argparse
import html
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as exc:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalar con: pip install openpyxl") from exc

try:
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTContainer, LTTextLine
except Exception as exc:  # pragma: no cover
    raise SystemExit("Falta pdfminer.six. Instalar con: pip install pdfminer.six") from exc


CHECKLIST_PREGUNTAS = [
    "1: Verificar en las aberturas las Tipologias/Medidas/Cantidades en MDT",
    "2: Verificar si completo todas las casillas de la hoja Control",
    "3: Verificar los Bipuntos, donde para cada Puerta Ventana Corrediza MODENA de 2 o 3 hojas, van 2 Bipuntos, si es de 4 o 6 hojas, 3 Bipuntos (25000 c/u), R60 YA INCLUYE BIPUNTOS.",
    "4: Si contiene aberturas de Alta Prestacion (R60), verificar e incluir despercicios entre 900 y 2100mm",
    "5: Verificar que coincida el PRECIO NETO con el PDF, el nombre de cliente y obra deseado, y pasar el PRECIO FINAL (+21% iva y dividir por 1.15 de descuento)",
    "6: Si hay variables chicas, solo hacer presupuesto y marcar diferencia, si hay cambios grandes (se suman aberturas, se pasa a alta prestacion, se pasa a DVH), rehacer cotizacion como alternativa 2.",
    "7: Avisar a Francisco y cargar en ODOO CRM con precio final y variables que solicite.",
    "8: Volver a comparar hoja Control con la Cotizacion y las aberturas con lo que fue solicitado el presupuesto",
]

CHECK_EMPTY = "\u2610"
CHECK_OK = "\u2611"
CHECK_NO = "\u2612"

GLASS_LAMIN = "Vidrios LAMIN"
GLASS_CRUDOS = "Vidrios Crudos"
GLASS_VIDRIAL = "Vidrios Vidrial"
GLASS_ORDER = [GLASS_LAMIN, GLASS_CRUDOS, GLASS_VIDRIAL]
PV_DETAIL_HEADERS = ["TIP.", "LADO1 VIDRIO", "LADO2 VIDRIO", "CANT.", "VIDRIO1", "VIDRIO2", "CAMARA"]
BIPUNTO_UNITARIO = 25000
RETAZO_USEFUL_MIN_DEFAULT = 1000
# Excepciones por codigo de perfil para retazos utiles.
# Formato: "CODIGO_SIN_MT": minimo_mm
# Si un perfil no esta aca, usa el minimo general de 1000 mm.
RETAZO_USEFUL_MIN_OVERRIDES: Dict[str, int] = {
    "0006": 1000,
    "0200": 1358,
    "0201": 1000,
    "0203": 1000,
    "0204": 1000,
    "0206": 2000,
    "0207": 1000,
    "0209": 1000,
    "0213": 1500,
    "0214": 2046,
    "0216": 1500,
    "0217": 2000,
    "0218": 650,
    "0219": 650,
    "0221": 1500,
    "0224": 1500,
    "0226": 2000,
    "0228": 1500,
    "0229": 1000,
    "0230": 2000,
    "0231": 2000,
    "0232": 2000,
    "0235": 1000,
    "0236": 1000,
    "0237": 2000,
    "0240": 3154,
    "0241": 2172,
    "0246": 1721,
    "0248": 1000,
    "0249": 1000,
    "0250": 1100,
    "0252": 1000,
    "0255": 1000,
    "0259": 1000,
    "0260": 1000,
    "0261": 2000,
    "0280": 1600,
    "0284": 2500,
    "0315": 2200,
    "0324": 2300,
    "0325": 1200,
    "0326": 1500,
    "0383": 1200,
}
# Excepciones de largo alternativo de barra para recalcular el retazo real.
# Formato: "CODIGO_SIN_MT": largo_barra_alternativa_mm
RETAZO_ALT_BAR_LENGTH_OVERRIDES: Dict[str, int] = {
    "0240": 4200,
}
ALTA_PRESTACION_RETAZO_RANGE_DEFAULT = (1000, 6000)
# Excepciones por codigo de perfil de Alta Prestacion.
# Formato: "CODIGO_SIN_MT": (min_mm, max_mm)
ALTA_PRESTACION_RETAZO_RANGE_OVERRIDES: Dict[str, tuple[int, int]] = {}
ALTA_PRESTACION_CODES = {
    "0326",
    "0397",
    "0402",
    "0619",
    "0621",
    "6501",
    "6502",
    "6503",
    "6504",
    "6505",
    "6506",
    "6507",
    "6508",
    "6510",
    "6511",
    "6512",
    "6513",
    "6514",
    "6515",
    "6516",
    "6517",
    "6520",
    "6521",
    "6522",
    "6523",
    "6524",
    "6526",
    "6528",
    "6529",
    "6530",
    "6531",
    "6532",
    "6533",
    "6534",
    "6535",
    "6536",
    "6537",
    "6538",
}

COLOR_MAP = {
    "N": "Negro",
    "NEGRO": "Negro",
    "B": "Blanco",
    "BLANCO": "Blanco",
    "A": "Anodizado",
    "ANODIZADO": "Anodizado",
}


@dataclass
class Cell:
    page: int
    x: float
    y: float
    text: str


@dataclass
class Row:
    page: int
    y: float
    cells: List[Cell]


def _normalize(text: str) -> str:
    clean = unicodedata.normalize("NFKD", text or "")
    clean = "".join(ch for ch in clean if not unicodedata.combining(ch))
    clean = clean.lower().replace("\xa0", " ")
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean


def _iter_text_lines(layout_obj: object) -> Iterable[LTTextLine]:
    if isinstance(layout_obj, LTTextLine):
        yield layout_obj
        return
    if isinstance(layout_obj, LTContainer):
        for child in layout_obj:
            yield from _iter_text_lines(child)


def _extract_cells(pdf_path: Path) -> List[Cell]:
    cells: List[Cell] = []
    for page_idx, page_layout in enumerate(extract_pages(str(pdf_path)), start=1):
        for line in _iter_text_lines(page_layout):
            text = re.sub(r"\s+", " ", line.get_text().strip())
            if not text:
                continue
            x0, y0, _, _ = line.bbox
            cells.append(Cell(page=page_idx, x=float(x0), y=float(y0), text=text))
    return cells


def _group_rows(cells: List[Cell], y_tolerance: float = 1.2) -> List[Row]:
    ordered = sorted(cells, key=lambda c: (c.page, -c.y, c.x))
    grouped: List[Row] = []

    current_cells: List[Cell] = []
    current_page: int | None = None
    current_y: float | None = None

    for cell in ordered:
        same_page = current_page == cell.page
        close_y = current_y is not None and abs(cell.y - current_y) <= y_tolerance
        if same_page and close_y:
            current_cells.append(cell)
            current_y = (current_y + cell.y) / 2.0
            continue

        if current_cells:
            grouped.append(
                Row(
                    page=current_page or 1,
                    y=current_y or 0.0,
                    cells=sorted(current_cells, key=lambda c: c.x),
                )
            )
        current_cells = [cell]
        current_page = cell.page
        current_y = cell.y

    if current_cells:
        grouped.append(
            Row(
                page=current_page or 1,
                y=current_y or 0.0,
                cells=sorted(current_cells, key=lambda c: c.x),
            )
        )
    return grouped


def _parse_number(raw: str) -> float | None:
    text = (raw or "").strip()
    if not text:
        return None
    text = text.replace("$", "").replace("\xa0", "").replace(" ", "")
    if "%" in text:
        return None
    if re.search(r"[^0-9,.\-]", text):
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        if text.count(",") == 1:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text and text.count(".") > 1:
        last = text.rfind(".")
        text = text[:last].replace(".", "") + text[last:]

    try:
        return float(text)
    except ValueError:
        return None


def _row_numeric_values(row: Row) -> List[float]:
    values: List[float] = []
    for cell in row.cells:
        value = _parse_number(cell.text)
        if value is not None:
            values.append(value)
    return values


def _prompt_color(arg_color: str | None) -> str:
    if arg_color:
        parsed = COLOR_MAP.get(_normalize(arg_color).upper())
        if parsed:
            return parsed

    prompt = "Color de la obra? [N]egro / [B]lanco / [A]nodizado: "
    while True:
        choice = input(prompt).strip()
        parsed = COLOR_MAP.get(_normalize(choice).upper())
        if parsed:
            return parsed
        print("Valor invalido. Usar N, B o A.")


def _parse_yes_no(value: str | None) -> bool | None:
    if value is None:
        return None
    token = _normalize(value).upper()
    if token in {"S", "SI", "Y", "YES", "1", "TRUE"}:
        return True
    if token in {"N", "NO", "0", "FALSE"}:
        return False
    return None


def _prompt_yes_no(arg_value: str | None, question: str) -> bool:
    parsed = _parse_yes_no(arg_value)
    if parsed is not None:
        return parsed

    while True:
        answer = input(question).strip()
        parsed = _parse_yes_no(answer)
        if parsed is not None:
            return parsed
        print("Valor invalido. Responder S o N.")


def _parse_non_negative_int(value: str | None) -> int | None:
    if value is None:
        return None
    txt = value.strip()
    if not txt:
        return None
    if not re.fullmatch(r"\d+", txt):
        return None
    return int(txt)


def _prompt_non_negative_int(arg_value: str | None, question: str) -> int:
    parsed = _parse_non_negative_int(arg_value)
    if parsed is not None:
        return parsed
    while True:
        answer = input(question).strip()
        parsed = _parse_non_negative_int(answer)
        if parsed is not None:
            return parsed
        print("Valor invalido. Ingresar un numero entero >= 0.")


def _extract_pp_profiles(pp_pdf: Path, color: str) -> List[Dict[str, object]]:
    rows = _group_rows(_extract_cells(pp_pdf))
    quantities_by_code: Dict[int, int] = {}

    for row in rows:
        code: int | None = None
        for cell in row.cells:
            match = re.search(r"\bMT-(\d{4})\b", cell.text, flags=re.IGNORECASE)
            if match:
                code = int(match.group(1))
                break
        if code is None:
            continue

        qty: int | None = None
        for cell in reversed(row.cells):
            parsed = _parse_number(cell.text)
            if parsed is None:
                continue
            if abs(parsed - round(parsed)) < 1e-9 and parsed >= 0:
                qty = int(round(parsed))
                break
        if qty is None:
            continue

        quantities_by_code[code] = quantities_by_code.get(code, 0) + qty

    result = [
        {"CODIGO": code, "COLOR": color, "CANTIDAD": qty}
        for code, qty in sorted(quantities_by_code.items(), key=lambda kv: kv[0])
    ]
    return result


def _build_retazo_rows(collected: Dict[tuple[int, str, int], int]) -> List[Dict[str, object]]:
    return [
        {
            "CODIGO": code,
            "COLOR": row_color,
            "LARGO_RETAZO": retazo,
            "CANTIDAD": qty,
        }
        for (code, row_color, retazo), qty in sorted(
            collected.items(), key=lambda item: (item[0][0], item[0][2], item[0][1])
        )
    ]


def _adjust_retazo_for_alt_bar_length(
    code_4: str,
    retazo_value: int,
    std_bar_length: int | None,
) -> int:
    alt_bar_length = RETAZO_ALT_BAR_LENGTH_OVERRIDES.get(code_4)
    if alt_bar_length is None or std_bar_length is None or std_bar_length <= 0:
        return retazo_value

    consumo_total = std_bar_length - retazo_value
    if consumo_total < 0:
        return retazo_value
    if consumo_total <= alt_bar_length:
        adjusted_retazo = alt_bar_length - consumo_total
        if adjusted_retazo >= 0:
            return adjusted_retazo
    return retazo_value


def _extract_split_retazos(opti_path: Path, color: str) -> tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    try:
        raw_html = opti_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raw_html = opti_path.read_text(encoding="latin-1")

    content = html.unescape(raw_html)
    blocks = list(
        re.finditer(
            r"Perfil\s+MT-(\d{4})(.*?)(?=Perfil\s+MT-\d{4}|$)",
            content,
            flags=re.IGNORECASE | re.DOTALL,
        )
    )

    collected_resto: Dict[tuple[int, str, int], int] = {}
    collected_alta: Dict[tuple[int, str, int], int] = {}

    for block in blocks:
        code_4 = block.group(1)
        chunk = block.group(2)
        qty_match = re.search(
            r"Cantidad\s+de\s+perfiles\s+utilizados\s+(\d+)(?:\s+de\s+(\d+))?",
            chunk,
            flags=re.IGNORECASE,
        )
        if not qty_match:
            continue
        qty = int(qty_match.group(1))
        if qty <= 0:
            continue
        std_bar_length = int(qty_match.group(2)) if qty_match.group(2) else None

        retazo_matches = re.finditer(r"\b[RD]\s*=\s*(\d+(?:[.,]\d+)?)", chunk, flags=re.IGNORECASE)
        retazo_value: int | None = None
        for retazo_match in retazo_matches:
            parsed = _parse_number(retazo_match.group(1))
            if parsed is None:
                continue
            rounded = _round_to_int(parsed)
            effective_retazo = _adjust_retazo_for_alt_bar_length(code_4, rounded, std_bar_length)
            if code_4 in ALTA_PRESTACION_CODES:
                min_retazo, max_retazo = ALTA_PRESTACION_RETAZO_RANGE_OVERRIDES.get(
                    code_4, ALTA_PRESTACION_RETAZO_RANGE_DEFAULT
                )
                if min_retazo <= effective_retazo <= max_retazo:
                    retazo_value = effective_retazo
                    break
            else:
                min_retazo = RETAZO_USEFUL_MIN_OVERRIDES.get(code_4, RETAZO_USEFUL_MIN_DEFAULT)
                if effective_retazo >= min_retazo:
                    retazo_value = effective_retazo
                    break
        if retazo_value is None:
            continue

        code = int(code_4)
        key = (code, color, retazo_value)
        if code_4 in ALTA_PRESTACION_CODES:
            collected_alta[key] = collected_alta.get(key, 0) + qty
        else:
            collected_resto[key] = collected_resto.get(key, 0) + qty

    return _build_retazo_rows(collected_resto), _build_retazo_rows(collected_alta)


def _classify_glass(glass_name: str) -> str:
    n = _normalize(glass_name)
    if any(token in n for token in ("templado", "camara", "cortado", "vidrial")):
        return GLASS_VIDRIAL
    if "laminado" in n:
        return GLASS_LAMIN
    if "float" in n or "pacifico" in n:
        return GLASS_CRUDOS
    return GLASS_VIDRIAL


def _extract_pv_summary(pv_pdf: Path) -> Dict[str, float]:
    rows = _group_rows(_extract_cells(pv_pdf))
    header_idx = -1
    vidrio_x = None
    medidas_x = None
    cantidad_x = None
    precio_x = None

    for idx, row in enumerate(rows):
        normalized_cells = {_normalize(c.text): c for c in row.cells}
        if "vidrio" in normalized_cells and "precio" in normalized_cells:
            header_idx = idx
            vidrio_x = normalized_cells["vidrio"].x
            precio_x = normalized_cells["precio"].x
            if "medidas" in normalized_cells:
                medidas_x = normalized_cells["medidas"].x
            if "cantidad" in normalized_cells:
                cantidad_x = normalized_cells["cantidad"].x
            break

    if header_idx < 0 or vidrio_x is None or precio_x is None:
        raise RuntimeError("No se encontro la tabla principal en PV.pdf")

    totals = {name: 0.0 for name in GLASS_ORDER}
    pending_glass_parts: List[str] = []
    header_tokens = {"tipologia", "vidrio", "medidas", "cantidad", "precio"}

    for row in rows[header_idx + 1 :]:
        row_text = _normalize(" ".join(cell.text for cell in row.cells))
        if row_text.startswith("lugar de entrega") or row_text.startswith("fecha de entrega") or row_text.startswith("pagina"):
            break

        price_value: float | None = None
        for cell in row.cells:
            if cell.x < (precio_x - 20):
                continue
            parsed = _parse_number(cell.text)
            if parsed is not None:
                price_value = parsed
                break

        # Algunos PDFs pegan Cantidad+Precio en una sola celda (ej. "6 1083434.00")
        # y esa celda queda desplazada hacia la columna Cantidad.
        if price_value is None and cantidad_x is not None:
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 20):
                    # Si ya cae en zona precio, el loop anterior debio tomarlo.
                    continue
                nums = re.findall(r"\d+(?:[.,]\d+)?", cell.text or "")
                if len(nums) < 2:
                    continue
                parsed_last = _parse_number(nums[-1])
                if parsed_last is not None:
                    price_value = parsed_last
                    break

        # Captura texto de la columna Vidrio con tolerancia mayor para casos de quiebre de linea.
        glass_text_parts: List[str] = []
        glass_left_limit = vidrio_x - 5
        glass_right_limit = (medidas_x - 20) if medidas_x is not None else (precio_x - 5)
        for cell in row.cells:
            if cell.x < glass_left_limit:
                continue
            if cell.x >= glass_right_limit:
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalpha() for ch in cell.text):
                glass_text_parts.append(cell.text)

        if glass_text_parts:
            pending_glass_parts.extend(glass_text_parts)

        if price_value is None:
            continue

        glass_name = " ".join(part for part in pending_glass_parts if str(part).strip()).strip()
        if not glass_name:
            # Fallback por si la descripcion cayo fuera del rango de columna
            raw_candidates = []
            for cell in row.cells:
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalpha() for ch in cell.text):
                    raw_candidates.append(cell.text)
            glass_name = " ".join(raw_candidates).strip()

        if not glass_name:
            continue

        category = _classify_glass(glass_name)
        totals[category] += float(price_value)
        pending_glass_parts = []

    return totals


def _extract_measure_sides(measures_text: str) -> tuple[int | None, int | None]:
    match = re.search(r"(\d+)\s*x\s*(\d+)", measures_text or "", flags=re.IGNORECASE)
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _map_pv_glass_code(glass_text: str) -> str:
    raw = re.sub(r"\s+", " ", glass_text or "").strip()
    if not raw:
        return ""

    normalized = _normalize(raw).replace(".", "")
    normalized = re.sub(r"(\d+)\s+mm\b", r"\1mm", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if "lamin op 3+3 cort" in normalized:
        return "LAMIN3+3OP"
    if "lamin inc 3+3 cortado dvh" in normalized:
        return "LAMIN3+3"
    if "float 4 c vidrial" in normalized:
        return "Float4"
    if "templado 10mm inc" in normalized or "templado de 10mm inc" in normalized:
        return "T10"
    if "templado 8mm inc" in normalized or "templado de 8mm inc" in normalized:
        return "T8"
    if "templado 6mm inc" in normalized or "templado de 6mm inc" in normalized:
        return "T6"
    if "templado 5mm inc" in normalized or "templado de 5mm inc" in normalized:
        return "T5"
    if "templado 4mm inc" in normalized or "templado de 4mm inc" in normalized:
        return "T4"
    if "templado solar neutro 6mm" in normalized:
        return "SN6mm"
    return ""


def _split_pv_glass_components(glass_text: str) -> tuple[str, str, int | None]:
    compact = re.sub(r"\s+", " ", glass_text or "").strip()
    if not compact:
        return "", "", None

    match = re.search(r"^(.*?)\bc[aá]mara\s+de\s+(\d+)\s*(.*)$", compact, flags=re.IGNORECASE)
    if not match:
        return compact, "", None

    left = match.group(1).strip(" -")
    right = match.group(3).strip(" -")
    if left and not right:
        right = left
    return left, right, int(match.group(2))


def _extract_pv_detailed_rows(pv_pdf: Path) -> List[Dict[str, object]]:
    rows = _group_rows(_extract_cells(pv_pdf))
    header_idx = -1
    tipologia_x = None
    vidrio_x = None
    medidas_x = None
    cantidad_x = None
    precio_x = None

    for idx, row in enumerate(rows):
        normalized_cells = {_normalize(c.text): c for c in row.cells}
        if "tipologia" in normalized_cells and "vidrio" in normalized_cells and "precio" in normalized_cells:
            header_idx = idx
            tipologia_x = normalized_cells["tipologia"].x
            vidrio_x = normalized_cells["vidrio"].x
            precio_x = normalized_cells["precio"].x
            if "medidas" in normalized_cells:
                medidas_x = normalized_cells["medidas"].x
            if "cantidad" in normalized_cells:
                cantidad_x = normalized_cells["cantidad"].x
            break

    if header_idx < 0 or tipologia_x is None or vidrio_x is None or precio_x is None:
        raise RuntimeError("No se encontro la tabla principal para detalle en PV.pdf")

    header_tokens = {"tipologia", "vidrio", "medidas", "cantidad", "precio"}
    pending_tip = ""
    pending_glass_parts: List[str] = []
    pending_measures = ""
    pending_qty: int | None = None
    detail_rows: List[Dict[str, object]] = []

    for row in rows[header_idx + 1 :]:
        row_text = _normalize(" ".join(cell.text for cell in row.cells))
        if row_text.startswith("lugar de entrega") or row_text.startswith("fecha de entrega") or row_text.startswith("pagina"):
            break

        tip_parts: List[str] = []
        for cell in row.cells:
            if cell.x < (tipologia_x - 20):
                continue
            if cell.x >= (vidrio_x - 5):
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalnum() for ch in cell.text):
                tip_parts.append(cell.text)
        if tip_parts:
            pending_tip = " ".join(tip_parts).strip()

        measures_right_limit = (cantidad_x - 5) if cantidad_x is not None else (precio_x - 5)
        measures_parts: List[str] = []
        if medidas_x is not None:
            for cell in row.cells:
                if cell.x < (medidas_x - 20):
                    continue
                if cell.x >= measures_right_limit:
                    continue
                if re.search(r"\d+\s*x\s*\d+", cell.text or "", flags=re.IGNORECASE):
                    measures_parts.append(cell.text)
        if measures_parts:
            pending_measures = " ".join(measures_parts).strip()
        elif not pending_measures:
            measure_match = re.search(r"\b\d+\s*x\s*\d+\b", " ".join(cell.text for cell in row.cells), flags=re.IGNORECASE)
            if measure_match:
                pending_measures = measure_match.group(0)

        if cantidad_x is not None:
            qty_value: int | None = None
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 5):
                    continue
                nums = re.findall(r"\d+", cell.text or "")
                if not nums:
                    continue
                qty_value = int(nums[0])
                break
            if qty_value is not None:
                pending_qty = qty_value

        glass_text_parts: List[str] = []
        glass_left_limit = vidrio_x - 5
        glass_right_limit = (medidas_x - 20) if medidas_x is not None else (precio_x - 5)
        for cell in row.cells:
            if cell.x < glass_left_limit:
                continue
            if cell.x >= glass_right_limit:
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalpha() for ch in cell.text):
                glass_text_parts.append(cell.text)
        if glass_text_parts:
            pending_glass_parts.extend(glass_text_parts)

        price_value: float | None = None
        for cell in row.cells:
            if cell.x < (precio_x - 20):
                continue
            parsed = _parse_number(cell.text)
            if parsed is not None:
                price_value = parsed
                break

        if price_value is None and cantidad_x is not None:
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 20):
                    continue
                nums = re.findall(r"\d+(?:[.,]\d+)?", cell.text or "")
                if len(nums) < 2:
                    continue
                parsed_last = _parse_number(nums[-1])
                if parsed_last is not None:
                    price_value = parsed_last
                    break

        if price_value is None:
            continue

        glass_name = " ".join(part for part in pending_glass_parts if str(part).strip()).strip()
        if not glass_name:
            raw_candidates = []
            for cell in row.cells:
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalpha() for ch in cell.text):
                    raw_candidates.append(cell.text)
            glass_name = " ".join(raw_candidates).strip()
        if not glass_name:
            pending_tip = ""
            pending_measures = ""
            pending_qty = None
            pending_glass_parts = []
            continue

        if not pending_tip:
            for cell in row.cells:
                if cell.x >= (vidrio_x - 5):
                    continue
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalnum() for ch in cell.text):
                    pending_tip = cell.text.strip()
                    break

        lado1, lado2 = _extract_measure_sides(pending_measures or glass_name)
        vidrio1_raw, vidrio2_raw, camara = _split_pv_glass_components(glass_name)
        if camara is None:
            pending_tip = ""
            pending_measures = ""
            pending_qty = None
            pending_glass_parts = []
            continue

        vidrio1 = _map_pv_glass_code(vidrio1_raw)
        vidrio2 = _map_pv_glass_code(vidrio2_raw)

        detail_rows.append(
            {
                "TIP.": pending_tip,
                "LADO1 VIDRIO": lado1,
                "LADO2 VIDRIO": lado2,
                "CANT.": pending_qty,
                "VIDRIO1": vidrio1,
                "VIDRIO2": vidrio2,
                "CAMARA": camara,
            }
        )

        pending_tip = ""
        pending_measures = ""
        pending_qty = None
        pending_glass_parts = []

    return detail_rows


def _find_row_by_label(rows: List[Row], label: str) -> Row:
    needle = _normalize(label).rstrip(":")
    for row in rows:
        for cell in row.cells:
            if _normalize(cell.text).startswith(needle):
                return row
    raise RuntimeError(f"No se encontro la fila '{label}' en INFORME.pdf")


def _extract_informe_summary(informe_pdf: Path) -> Dict[str, float | None]:
    rows = _group_rows(_extract_cells(informe_pdf))

    accesorios_row = _find_row_by_label(rows, "Accesorios en Pesos")
    vidrios_row = _find_row_by_label(rows, "Vidrios en Pesos")

    accesorios_vals = _row_numeric_values(accesorios_row)
    vidrios_vals = _row_numeric_values(vidrios_row)

    if len(accesorios_vals) < 1:
        raise RuntimeError("No se pudo leer Consumo de 'Accesorios en Pesos'")
    if len(vidrios_vals) < 1:
        raise RuntimeError("No se pudo leer Consumo de 'Vidrios en Pesos'")

    return {
        "kg_consumidos": None,
        "kg_comprados": None,
        "accesorios_resto": accesorios_vals[0],
        "vidrios_pesos_consumo": vidrios_vals[0],
    }


def _round_to_int(value: float) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _write_integer(ws, row: int, col: int, value: float) -> None:
    cell = ws.cell(row=row, column=col, value=_round_to_int(float(value)))
    cell.number_format = "0"


def _write_integer_or_blank(ws, row: int, col: int, value: float | None) -> None:
    if value is None:
        ws.cell(row=row, column=col, value="")
        return
    _write_integer(ws, row, col, value)


def _auto_width(ws, min_width: int = 12, max_width: int = 95) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _write_excel(
    output_path: Path,
    perfiles: List[Dict[str, object]],
    retazos_resto: List[Dict[str, object]],
    retazos_alta_prestacion: List[Dict[str, object]],
    pv_detail_rows: List[Dict[str, object]],
    incluir_retazos: bool,
    bipuntos_cantidad: int,
    pv_totals: Dict[str, float],
    informe_summary: Dict[str, float | None],
) -> Dict[str, float | str]:
    total_vidrios_pv = sum(pv_totals.get(k, 0.0) for k in GLASS_ORDER)
    total_vidrios_informe = float(informe_summary["vidrios_pesos_consumo"])
    total_vidrios_pv_int = _round_to_int(total_vidrios_pv)
    total_vidrios_informe_int = _round_to_int(total_vidrios_informe)
    diferencia = total_vidrios_pv_int - total_vidrios_informe_int
    coincide = diferencia == 0
    estado = "OK" if coincide else f"NO COINCIDE (dif: {diferencia:+d})"

    wb = Workbook()
    ws = wb.active
    ws.title = "Control"
    bold = Font(bold=True)

    row = 1
    ws.cell(row=row, column=1, value="Perfiles (PP.pdf)").font = bold
    row += 1
    for col, header in enumerate(["CODIGO", "COLOR", "CANTIDAD"], start=1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    for item in perfiles:
        ws.cell(row=row, column=1, value=int(item["CODIGO"]))
        ws.cell(row=row, column=2, value=str(item["COLOR"]))
        ws.cell(row=row, column=3, value=int(item["CANTIDAD"]))
        row += 1

    if incluir_retazos:
        row += 1
        ws.cell(row=row, column=1, value="Retazos Utiles - Resto (Opti.htm)").font = bold
        row += 1
        for col, header in enumerate(["CODIGO", "COLOR", "LARGO RETAZO", "CANTIDAD"], start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1

        for item in retazos_resto:
            ws.cell(row=row, column=1, value=int(item["CODIGO"]))
            ws.cell(row=row, column=2, value=str(item["COLOR"]))
            ws.cell(row=row, column=3, value=int(item["LARGO_RETAZO"]))
            ws.cell(row=row, column=4, value=int(item["CANTIDAD"]))
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Retazos Alta Prestacion (Opti.htm)").font = bold
        row += 1
        for col, header in enumerate(["CODIGO", "COLOR", "LARGO RETAZO", "CANTIDAD"], start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1

        for item in retazos_alta_prestacion:
            ws.cell(row=row, column=1, value=int(item["CODIGO"]))
            ws.cell(row=row, column=2, value=str(item["COLOR"]))
            ws.cell(row=row, column=3, value=int(item["LARGO_RETAZO"]))
            ws.cell(row=row, column=4, value=int(item["CANTIDAD"]))
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Detalle Vidrios (PV.pdf)").font = bold
    row += 1
    for col, header in enumerate(PV_DETAIL_HEADERS, start=1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    for item in pv_detail_rows:
        for col, header in enumerate(PV_DETAIL_HEADERS, start=1):
            value = item.get(header, "")
            ws.cell(row=row, column=col, value="" if value is None else value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Resumen Vidrios (PV.pdf)").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Categoria").font = bold
    ws.cell(row=row, column=2, value="Precio").font = bold
    row += 1

    for category in GLASS_ORDER:
        ws.cell(row=row, column=1, value=category)
        _write_integer(ws, row, 2, float(pv_totals.get(category, 0.0)))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Resumen Informe (INFORME.pdf)").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Concepto").font = bold
    ws.cell(row=row, column=2, value="Valor").font = bold
    row += 1

    ws.cell(row=row, column=1, value="Accesorios resto")
    _write_integer(ws, row, 2, float(informe_summary["accesorios_resto"]))
    row += 1

    ws.cell(row=row, column=1, value="Bipuntos")
    _write_integer(ws, row, 2, float(bipuntos_cantidad * BIPUNTO_UNITARIO))
    row += 1

    ws.cell(row=row, column=1, value="Kg consumidos")
    _write_integer_or_blank(ws, row, 2, informe_summary["kg_consumidos"])
    row += 1

    ws.cell(row=row, column=1, value="Kg comprados")
    _write_integer_or_blank(ws, row, 2, informe_summary["kg_comprados"])
    row += 1

    ws.cell(row=row, column=1, value="Vidrios en Pesos (INFORME - Consumo)")
    _write_integer(ws, row, 2, total_vidrios_informe)
    row += 1

    ws.cell(row=row, column=1, value="Total Vidrios (PV)")
    _write_integer(ws, row, 2, total_vidrios_pv)
    row += 1

    ws.cell(row=row, column=1, value="Chequeo Vidrios (PV vs INFORME)")
    ws.cell(row=row, column=2, value=estado)

    checklist = wb.create_sheet("Checklist")
    checklist.cell(row=1, column=1, value="Pregunta").font = bold
    checklist.cell(row=1, column=2, value="Chequeado").font = bold

    check_validation = DataValidation(
        type="list",
        formula1=f'"{CHECK_EMPTY},{CHECK_OK},{CHECK_NO}"',
        allow_blank=False,
    )
    checklist.add_data_validation(check_validation)

    for idx, pregunta in enumerate(CHECKLIST_PREGUNTAS, start=2):
        checklist.cell(row=idx, column=1, value=pregunta)
        check_cell = checklist.cell(row=idx, column=2, value=CHECK_EMPTY)
        check_cell.alignment = Alignment(horizontal="center", vertical="center")
        check_validation.add(check_cell)

    _auto_width(ws)
    _auto_width(checklist, min_width=14, max_width=130)
    wb.save(output_path)

    return {
        "total_vidrios_pv": total_vidrios_pv_int,
        "total_vidrios_informe": total_vidrios_informe_int,
        "diferencia": diferencia,
        "estado": estado,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Control de cotizaciones: perfiles (PP), vidrios (PV) e informe final."
    )
    parser.add_argument("--pp", type=Path, default=Path("PP.pdf"), help="Ruta a PP.pdf")
    parser.add_argument("--pv", type=Path, default=Path("PV.pdf"), help="Ruta a PV.pdf")
    parser.add_argument("--informe", type=Path, default=Path("INFORME.pdf"), help="Ruta a INFORME.pdf")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("Control_Cotizacion.xlsx"),
        help="Excel de salida",
    )
    parser.add_argument(
        "--color",
        type=str,
        default=None,
        help="Color de obra: N/B/A o nombre completo",
    )
    parser.add_argument(
        "--alta",
        type=str,
        default=None,
        help="Compatibilidad hacia atras. Usar --retazos.",
    )
    parser.add_argument(
        "--retazos",
        type=str,
        default=None,
        help="Analizar retazos utiles desde Opti.htm? S/N",
    )
    parser.add_argument(
        "--opti",
        type=Path,
        default=Path(r"C:\MDT_Winproject_2\Proyecto\Opti.htm"),
        help="Ruta al archivo Opti.htm para retazos utiles",
    )
    parser.add_argument(
        "--bipuntos",
        type=str,
        default=None,
        help="Cantidad de bipuntos (entero >= 0)",
    )
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    color = _prompt_color(args.color)
    bipuntos = _prompt_non_negative_int(
        args.bipuntos,
        (
            "Cuantos Bipuntos tiene la obra?\n"
            "Corredizas de 2 y 3 hojas: 2 bipuntos\n"
            "Corredizas de 4 y 6 hojas: 3 bipuntos\n"
            "Las corredizas de alta prestacion YA INCLUYEN y no debe contarlas nuevamente.\n"
            "Ingrese cantidad total de bipuntos: "
        ),
    )
    analizar_retazos = _prompt_yes_no(
        args.retazos if args.retazos is not None else args.alta,
        "Analizar retazos utiles desde Opti.htm? [S/N]: ",
    )

    for path in (args.pp, args.pv, args.informe):
        if not path.exists():
            raise SystemExit(f"No existe el archivo: {path}")

    retazos_resto: List[Dict[str, object]] = []
    retazos_alta_prestacion: List[Dict[str, object]] = []
    if analizar_retazos:
        if not args.opti.exists():
            raise SystemExit(f"No existe el archivo Opti.htm en: {args.opti}")
        retazos_resto, retazos_alta_prestacion = _extract_split_retazos(args.opti, color)

    perfiles = _extract_pp_profiles(args.pp, color)
    if not perfiles:
        print("Aviso: no se encontraron perfiles MT-XXXX en PP.pdf", file=sys.stderr)

    pv_detail_rows = _extract_pv_detailed_rows(args.pv)
    pv_totals = _extract_pv_summary(args.pv)
    informe_summary = _extract_informe_summary(args.informe)
    resultado = _write_excel(
        args.out,
        perfiles,
        retazos_resto,
        retazos_alta_prestacion,
        pv_detail_rows,
        analizar_retazos,
        bipuntos,
        pv_totals,
        informe_summary,
    )

    print(f"Excel generado: {args.out.resolve()}")
    print(f"Color aplicado a perfiles: {color}")
    print(f"Bipuntos: {bipuntos} (costo total: {bipuntos * BIPUNTO_UNITARIO})")
    if analizar_retazos:
        print(f"Retazos utiles - resto encontrados: {len(retazos_resto)}")
        print(f"Retazos Alta Prestacion encontrados: {len(retazos_alta_prestacion)}")
    else:
        print("Retazos: no se solicitaron")
    print(
        "Chequeo vidrios: "
        f"{resultado['estado']} | "
        f"PV={resultado['total_vidrios_pv']} vs INFORME={resultado['total_vidrios_informe']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
