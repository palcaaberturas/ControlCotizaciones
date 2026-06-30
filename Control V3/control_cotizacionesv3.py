from __future__ import annotations

import argparse
import html
import os
import re
import shutil
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as exc:  # pragma: no cover
    raise SystemExit("Falta openpyxl. Instalar con: pip install openpyxl") from exc

try:
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTContainer, LTTextLine
except Exception as exc:  # pragma: no cover
    raise SystemExit("Falta pdfminer.six. Instalar con: pip install pdfminer.six") from exc


CHECKLIST_PREGUNTAS = [
    "1: Verificar en las aberturas las Tipologias/Medidas/Cantidades en MDT",
    "2: Verificar si completo todas las casillas de la hoja Control",
    "3: Verificar los Bipuntos, donde para cada Puerta Ventana Corrediza MODENA de 2 o 3 hojas, van 2 Bipuntos, si es de 4 o 6 hojas, 3 Bipuntos (25000 c/u), R60 YA INCLUYE BIPUNTOS.",
    "4: Si contiene aberturas de Alta Prestacion (R60), verificar e incluir despercicios entre 900 y 2100mm",
    "5: Verificar que coincida el PRECIO NETO con el PDF, el nombre de cliente y obra deseado, y pasar el PRECIO FINAL (+21% iva y dividir por 1.15 de descuento)",
    "6: Si hay variables chicas, solo hacer presupuesto y marcar diferencia, si hay cambios grandes (se suman aberturas, se pasa a alta prestacion, se pasa a DVH), rehacer cotizacion como alternativa 2.",
    "7: Avisar a Francisco y cargar en ODOO CRM con precio final y variables que solicite.",
    "8: Volver a comparar hoja Control con la Cotizacion y las aberturas con lo que fue solicitado el presupuesto",
]

CHECK_EMPTY = "\u2610"
CHECK_OK = "\u2611"
CHECK_NO = "\u2612"

GLASS_LAMIN = "Vidrios LAMIN"
GLASS_CRUDOS = "Vidrios Crudos"
GLASS_DVH = "Vidrios Vidrial DVH"
GLASS_TEMPLADOS = "Vidrios Vidrial Templados"
GLASS_ORDER = [GLASS_LAMIN, GLASS_CRUDOS, GLASS_DVH, GLASS_TEMPLADOS]
PV_DETAIL_HEADERS = ["TIP.", "LADO1 VIDRIO", "LADO2 VIDRIO", "CANT.", "VIDRIO1", "VIDRIO2", "CAMARA"]
BIPUNTO_UNITARIO = 25000
RETAZO_USEFUL_MIN_DEFAULT = 1000
# Excepciones por codigo de perfil para retazos utiles.
# Formato: "CODIGO_SIN_MT": minimo_mm
# Si un perfil no esta aca, usa el minimo general de 1000 mm.
RETAZO_USEFUL_MIN_OVERRIDES: Dict[str, int] = {
    "0006": 1000,
    "0200": 1358,
    "0201": 1000,
    "0203": 1000,
    "0204": 1000,
    "0206": 2000,
    "0207": 1000,
    "0209": 1000,
    "0213": 1500,
    "0214": 2046,
    "0216": 1500,
    "0217": 2000,
    "0218": 650,
    "0219": 650,
    "0221": 1500,
    "0224": 1500,
    "0226": 2000,
    "0228": 1500,
    "0229": 1000,
    "0230": 2000,
    "0231": 2000,
    "0232": 2000,
    "0235": 1000,
    "0236": 1000,
    "0237": 2000,
    "0240": 3154,
    "0241": 2172,
    "0246": 1721,
    "0248": 1000,
    "0249": 1000,
    "0250": 1100,
    "0252": 1000,
    "0255": 1000,
    "0259": 1000,
    "0260": 1000,
    "0261": 2000,
    "0280": 1600,
    "0284": 2500,
    "0315": 2200,
    "0324": 2300,
    "0325": 1200,
    "0326": 1500,
    "0383": 1200,
}
# Excepciones de largo alternativo de barra para recalcular el retazo real.
# Formato: "CODIGO_SIN_MT": largo_barra_alternativa_mm
RETAZO_ALT_BAR_LENGTH_OVERRIDES: Dict[str, int] = {
    "0240": 4200,
}
ALTA_PRESTACION_RETAZO_RANGE_DEFAULT = (1250, 6200)
# Excepciones por codigo de perfil de Alta Prestacion.
# Formato: "CODIGO_SIN_MT": (min_mm, max_mm)
ALTA_PRESTACION_RETAZO_RANGE_OVERRIDES: Dict[str, tuple[int, int]] = {
    "0397": (1500, 6200),
    "0402": (2500, 6200),
    "0615": (1500, 6200),
    "0619": (1000, 6200),
    "0621": (1500, 6200),
    "6503": (2500, 6200),
    "6504": (2500, 6200),
    "6505": (2500, 6200),
    "6508": (1500, 6200),
    "6510": (2700, 6200),
    "6511": (2700, 6200),
    "6512": (3000, 6200),
    "6513": (3800, 6200),
    "6514": (2500, 6200),
    "6515": (4000, 6200),
    "6523": (2400, 6200),
    "6528": (2350, 6200),
    "6537": (1500, 6200),
    "6538": (3000, 6200),
}
ALTA_PRESTACION_CODES = {
    "0326",
    "0397",
    "0402",
    "0615",
    "0619",
    "0621",
    "6501",
    "6502",
    "6503",
    "6504",
    "6505",
    "6506",
    "6507",
    "6508",
    "6510",
    "6511",
    "6512",
    "6513",
    "6514",
    "6515",
    "6516",
    "6517",
    "6520",
    "6521",
    "6522",
    "6523",
    "6524",
    "6526",
    "6528",
    "6529",
    "6530",
    "6531",
    "6532",
    "6533",
    "6534",
    "6535",
    "6536",
    "6537",
    "6538",
    "4007", 
}

COLOR_MAP = {
    "N": "Negro",
    "NEGRO": "Negro",
    "B": "Blanco",
    "BLANCO": "Blanco",
    "A": "Anodizado",
    "ANODIZADO": "Anodizado",
}


@dataclass
class Cell:
    page: int
    x: float
    y: float
    text: str


@dataclass
class Row:
    page: int
    y: float
    cells: List[Cell]


def _normalize(text: str) -> str:
    clean = unicodedata.normalize("NFKD", text or "")
    clean = "".join(ch for ch in clean if not unicodedata.combining(ch))
    clean = clean.lower().replace("\xa0", " ")
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean

def _iter_text_lines(layout_obj: object) -> Iterable[LTTextLine]:
    if isinstance(layout_obj, LTTextLine):
        yield layout_obj
        return
    if isinstance(layout_obj, LTContainer):
        for child in layout_obj:
            yield from _iter_text_lines(child)

def _extract_cells(pdf_path: Path) -> List[Cell]:
    cells: List[Cell] = []
    for page_idx, page_layout in enumerate(extract_pages(str(pdf_path)), start=1):
        for line in _iter_text_lines(page_layout):
            text = re.sub(r"\s+", " ", line.get_text().strip())
            if not text:
                continue
            x0, y0, _, _ = line.bbox
            cells.append(Cell(page=page_idx, x=float(x0), y=float(y0), text=text))
    return cells

def _iter_text_lines(layout_obj: object) -> Iterable[LTTextLine]:
    if isinstance(layout_obj, LTTextLine):
        yield layout_obj
        return
    if isinstance(layout_obj, LTContainer):
        for child in layout_obj:
            yield from _iter_text_lines(child)


def _group_rows(cells: List[Cell], y_tolerance: float = 1.2) -> List[Row]:
    ordered = sorted(cells, key=lambda c: (c.page, -c.y, c.x))
    grouped: List[Row] = []

    current_cells: List[Cell] = []
    current_page: int | None = None
    current_y: float | None = None

    for cell in ordered:
        same_page = current_page == cell.page
        close_y = current_y is not None and abs(cell.y - current_y) <= y_tolerance
        if same_page and close_y:
            current_cells.append(cell)
            current_y = (current_y + cell.y) / 2.0
            continue

        if current_cells:
            grouped.append(
                Row(
                    page=current_page or 1,
                    y=current_y or 0.0,
                    cells=sorted(current_cells, key=lambda c: c.x),
                )
            )
        current_cells = [cell]
        current_page = cell.page
        current_y = cell.y

    if current_cells:
        grouped.append(
            Row(
                page=current_page or 1,
                y=current_y or 0.0,
                cells=sorted(current_cells, key=lambda c: c.x),
            )
        )
    return grouped


def _parse_number(raw: str) -> float | None:
    text = (raw or "").strip()
    if not text:
        return None
    text = text.replace("$", "").replace("\xa0", "").replace(" ", "")
    if "%" in text:
        return None
    if re.search(r"[^0-9,.\-]", text):
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        if text.count(",") == 1:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text and text.count(".") > 1:
        last = text.rfind(".")
        text = text[:last].replace(".", "") + text[last:]

    try:
        return float(text)
    except ValueError:
        return None


def _row_numeric_values(row: Row) -> List[float]:
    values: List[float] = []
    for cell in row.cells:
        value = _parse_number(cell.text)
        if value is not None:
            values.append(value)
    return values


def _prompt_color(arg_color: str | None) -> str:
    if arg_color:
        parsed = COLOR_MAP.get(_normalize(arg_color).upper())
        if parsed:
            return parsed

    prompt = "Color de la obra? [N]egro / [B]lanco / [A]nodizado: "
    while True:
        choice = input(prompt).strip()
        parsed = COLOR_MAP.get(_normalize(choice).upper())
        if parsed:
            return parsed
        print("Valor invalido. Usar N, B o A.")


def _parse_yes_no(value: str | None) -> bool | None:
    if value is None:
        return None
    token = _normalize(value).upper()
    if token in {"S", "SI", "Y", "YES", "1", "TRUE"}:
        return True
    if token in {"N", "NO", "0", "FALSE"}:
        return False
    return None


def _prompt_yes_no(arg_value: str | None, question: str) -> bool:
    parsed = _parse_yes_no(arg_value)
    if parsed is not None:
        return parsed

    while True:
        answer = input(question).strip()
        parsed = _parse_yes_no(answer)
        if parsed is not None:
            return parsed
        print("Valor invalido. Responder S o N.")


def _parse_non_negative_int(value: str | None) -> int | None:
    if value is None:
        return None
    txt = value.strip()
    if not txt:
        return None
    if not re.fullmatch(r"\d+", txt):
        return None
    return int(txt)


def _prompt_non_negative_int(arg_value: str | None, question: str) -> int:
    parsed = _parse_non_negative_int(arg_value)
    if parsed is not None:
        return parsed
    while True:
        answer = input(question).strip()
        parsed = _parse_non_negative_int(answer)
        if parsed is not None:
            return parsed
        print("Valor invalido. Ingresar un numero entero >= 0.")


def _build_retazo_rows(collected: Dict[tuple[int, str, int], int]) -> List[Dict[str, object]]:
    return [
        {
            "CODIGO": code,
            "COLOR": row_color,
            "LARGO_RETAZO": retazo,
            "CANTIDAD": qty,
        }
        for (code, row_color, retazo), qty in sorted(
            collected.items(), key=lambda item: (item[0][0], item[0][2], item[0][1])
        )
    ]


def _adjust_retazo_for_alt_bar_length(
    code_4: str,
    retazo_value: int,
    std_bar_length: int | None,
) -> int:
    alt_bar_length = RETAZO_ALT_BAR_LENGTH_OVERRIDES.get(code_4)
    if alt_bar_length is None or std_bar_length is None or std_bar_length <= 0:
        return retazo_value

    consumo_total = std_bar_length - retazo_value
    if consumo_total < 0:
        return retazo_value
    if consumo_total <= alt_bar_length:
        adjusted_retazo = alt_bar_length - consumo_total
        if adjusted_retazo >= 0:
            return adjusted_retazo
    return retazo_value


def _extract_split_retazos(opti_path: Path, color: str) -> tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    try:
        raw_html = opti_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raw_html = opti_path.read_text(encoding="latin-1")

    content = html.unescape(raw_html)
    blocks = list(
        re.finditer(
            r"Perfil\s+MT-(\d{4})(.*?)(?=Perfil\s+MT-\d{4}|$)",
            content,
            flags=re.IGNORECASE | re.DOTALL,
        )
    )

    collected_resto: Dict[tuple[int, str, int], int] = {}
    collected_alta: Dict[tuple[int, str, int], int] = {}

    for block in blocks:
        code_4 = block.group(1)
        chunk = block.group(2)
        qty_match = re.search(
            r"Cantidad\s+de\s+perfiles\s+utilizados\s+(\d+)(?:\s+de\s+(\d+))?",
            chunk,
            flags=re.IGNORECASE,
        )
        if not qty_match:
            continue
        qty = int(qty_match.group(1))
        if qty <= 0:
            continue
        std_bar_length = int(qty_match.group(2)) if qty_match.group(2) else None

        retazo_matches = re.finditer(r"\b[RD]\s*=\s*(\d+(?:[.,]\d+)?)", chunk, flags=re.IGNORECASE)
        retazo_value: int | None = None
        for retazo_match in retazo_matches:
            parsed = _parse_number(retazo_match.group(1))
            if parsed is None:
                continue
            rounded = _round_to_int(parsed)
            effective_retazo = _adjust_retazo_for_alt_bar_length(code_4, rounded, std_bar_length)
            if code_4 in ALTA_PRESTACION_CODES:
                min_retazo, max_retazo = ALTA_PRESTACION_RETAZO_RANGE_OVERRIDES.get(
                    code_4, ALTA_PRESTACION_RETAZO_RANGE_DEFAULT
                )
                if min_retazo <= effective_retazo <= max_retazo:
                    retazo_value = effective_retazo
                    break
            else:
                min_retazo = RETAZO_USEFUL_MIN_OVERRIDES.get(code_4, RETAZO_USEFUL_MIN_DEFAULT)
                if effective_retazo >= min_retazo:
                    retazo_value = effective_retazo
                    break
        if retazo_value is None:
            continue

        code = int(code_4)
        key = (code, color, retazo_value)
        if code_4 in ALTA_PRESTACION_CODES:
            collected_alta[key] = collected_alta.get(key, 0) + qty
        else:
            collected_resto[key] = collected_resto.get(key, 0) + qty

    return _build_retazo_rows(collected_resto), _build_retazo_rows(collected_alta)

def _extract_opti_profiles(opti_path: Path, color: str) -> List[Dict[str, object]]:
    """Extrae la suma total de perfiles a pedir directamente desde el archivo Opti.htm"""
    try:
        raw_html = opti_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raw_html = opti_path.read_text(encoding="latin-1")

    content = html.unescape(raw_html)
    
    # Encontrar todos los bloques de perfiles en el HTML
    blocks = re.finditer(
        r"Perfil\s+MT-(\d{4})(.*?)(?=Perfil\s+MT-\d{4}|$)",
        content,
        flags=re.IGNORECASE | re.DOTALL,
    )

    quantities_by_code: Dict[int, int] = {}

    for block in blocks:
        code = int(block.group(1))
        chunk = block.group(2)
        
        # Buscar la cantidad utilizada en este bloque específico
        qty_match = re.search(
            r"Cantidad\s+de\s+perfiles\s+utilizados\s+(\d+)",
            chunk,
            flags=re.IGNORECASE,
        )
        
        if qty_match:
            qty = int(qty_match.group(1))
            # Sumar la cantidad al acumulado de ese código de perfil
            quantities_by_code[code] = quantities_by_code.get(code, 0) + qty

    # Formatear la lista de diccionarios igual que lo hacía la función anterior
    result = [
        {"CODIGO": code, "COLOR": color, "CANTIDAD": qty}
        for code, qty in sorted(quantities_by_code.items(), key=lambda kv: kv[0])
    ]
    return result


def _classify_glass(glass_name: str) -> str:
    n = _normalize(glass_name)
    if "laminado" in n:
        return GLASS_LAMIN
    if "float" in n or "pacifico" in n:
        return GLASS_CRUDOS
    if "camara" in n:
        return GLASS_DVH
    return GLASS_TEMPLADOS


def _extract_pv_summary(pv_pdf: Path) -> Dict[str, float]:
    rows = _group_rows(_extract_cells(pv_pdf))
    header_idx = -1
    vidrio_x = None
    medidas_x = None
    cantidad_x = None
    precio_x = None

    for idx, row in enumerate(rows):
        normalized_cells = {_normalize(c.text): c for c in row.cells}
        if "vidrio" in normalized_cells and "precio" in normalized_cells:
            header_idx = idx
            vidrio_x = normalized_cells["vidrio"].x
            precio_x = normalized_cells["precio"].x
            if "medidas" in normalized_cells:
                medidas_x = normalized_cells["medidas"].x
            if "cantidad" in normalized_cells:
                cantidad_x = normalized_cells["cantidad"].x
            break

    if header_idx < 0 or vidrio_x is None or precio_x is None:
        raise RuntimeError("No se encontro la tabla principal en PV.pdf")

    totals = {name: 0.0 for name in GLASS_ORDER}
    pending_glass_parts: List[str] = []
    header_tokens = {"tipologia", "vidrio", "medidas", "cantidad", "precio"}

    for row in rows[header_idx + 1 :]:
        row_text = _normalize(" ".join(cell.text for cell in row.cells))
        if row_text.startswith("lugar de entrega") or row_text.startswith("fecha de entrega") or row_text.startswith("pagina"):
            break

        price_value: float | None = None
        for cell in row.cells:
            if cell.x < (precio_x - 20):
                continue
            parsed = _parse_number(cell.text)
            if parsed is not None:
                price_value = parsed
                break

        # Algunos PDFs pegan Cantidad+Precio en una sola celda (ej. "6 1083434.00")
        # y esa celda queda desplazada hacia la columna Cantidad.
        if price_value is None and cantidad_x is not None:
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 20):
                    # Si ya cae en zona precio, el loop anterior debio tomarlo.
                    continue
                nums = re.findall(r"\d+(?:[.,]\d+)?", cell.text or "")
                if len(nums) < 2:
                    continue
                parsed_last = _parse_number(nums[-1])
                if parsed_last is not None:
                    price_value = parsed_last
                    break

        # Captura texto de la columna Vidrio con tolerancia mayor para casos de quiebre de linea.
        glass_text_parts: List[str] = []
        glass_left_limit = vidrio_x - 5
        glass_right_limit = (medidas_x - 20) if medidas_x is not None else (precio_x - 5)
        for cell in row.cells:
            if cell.x < glass_left_limit:
                continue
            if cell.x >= glass_right_limit:
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalpha() for ch in cell.text):
                glass_text_parts.append(cell.text)

        if glass_text_parts:
            pending_glass_parts.extend(glass_text_parts)

        if price_value is None:
            continue

        glass_name = " ".join(part for part in pending_glass_parts if str(part).strip()).strip()
        if not glass_name:
            # Fallback por si la descripcion cayo fuera del rango de columna
            raw_candidates = []
            for cell in row.cells:
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalpha() for ch in cell.text):
                    raw_candidates.append(cell.text)
            glass_name = " ".join(raw_candidates).strip()

        if not glass_name:
            continue

        category = _classify_glass(glass_name)
        totals[category] += float(price_value)
        pending_glass_parts = []

    return totals


def _extract_measure_sides(measures_text: str) -> tuple[int | None, int | None]:
    match = re.search(r"(\d+)\s*x\s*(\d+)", measures_text or "", flags=re.IGNORECASE)
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _map_pv_glass_code(glass_text: str) -> str:
    raw = re.sub(r"\s+", " ", glass_text or "").strip()
    if not raw:
        return ""

    normalized = _normalize(raw).replace(".", "")
    normalized = re.sub(r"(\d+)\s+mm\b", r"\1mm", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if "lamin op 3+3 cort" in normalized:
        return "LAMIN3+3OP"
    if "lamin inc 3+3 cortado dvh" in normalized:
        return "LAMIN3+3"
    if "float 4 c vidrial" in normalized:
        return "Float4"
    if "templado 10mm inc" in normalized or "templado de 10mm inc" in normalized:
        return "T10"
    if "templado 8mm inc" in normalized or "templado de 8mm inc" in normalized:
        return "T8"
    if "templado 6mm inc" in normalized or "templado de 6mm inc" in normalized:
        return "T6"
    if "templado 5mm inc" in normalized or "templado de 5mm inc" in normalized:
        return "T5"
    if "templado 4mm inc" in normalized or "templado de 4mm inc" in normalized:
        return "T4"
    if "templado solar neutro 6mm" in normalized:
        return "SN6mm"
    return ""


def _split_pv_glass_components(glass_text: str) -> tuple[str, str, int | None]:
    compact = re.sub(r"\s+", " ", glass_text or "").strip()
    if not compact:
        return "", "", None

    match = re.search(r"^(.*?)\bc[aá]mara\s+de\s+(\d+)\s*(.*)$", compact, flags=re.IGNORECASE)
    if not match:
        return compact, "", None

    left = match.group(1).strip(" -")
    right = match.group(3).strip(" -")
    if left and not right:
        right = left
    return left, right, int(match.group(2))


def _extract_pv_detailed_rows(pv_pdf: Path) -> List[Dict[str, object]]:
    rows = _group_rows(_extract_cells(pv_pdf))
    header_idx = -1
    tipologia_x = None
    vidrio_x = None
    medidas_x = None
    cantidad_x = None
    precio_x = None

    for idx, row in enumerate(rows):
        normalized_cells = {_normalize(c.text): c for c in row.cells}
        if "tipologia" in normalized_cells and "vidrio" in normalized_cells and "precio" in normalized_cells:
            header_idx = idx
            tipologia_x = normalized_cells["tipologia"].x
            vidrio_x = normalized_cells["vidrio"].x
            precio_x = normalized_cells["precio"].x
            if "medidas" in normalized_cells:
                medidas_x = normalized_cells["medidas"].x
            if "cantidad" in normalized_cells:
                cantidad_x = normalized_cells["cantidad"].x
            break

    if header_idx < 0 or tipologia_x is None or vidrio_x is None or precio_x is None:
        raise RuntimeError("No se encontro la tabla principal para detalle en PV.pdf")

    header_tokens = {"tipologia", "vidrio", "medidas", "cantidad", "precio"}
    pending_tip = ""
    pending_glass_parts: List[str] = []
    pending_measures = ""
    pending_qty: int | None = None
    detail_rows: List[Dict[str, object]] = []

    for row in rows[header_idx + 1 :]:
        row_text = _normalize(" ".join(cell.text for cell in row.cells))
        if row_text.startswith("lugar de entrega") or row_text.startswith("fecha de entrega") or row_text.startswith("pagina"):
            break

        tip_parts: List[str] = []
        for cell in row.cells:
            if cell.x < (tipologia_x - 20):
                continue
            if cell.x >= (vidrio_x - 5):
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalnum() for ch in cell.text):
                tip_parts.append(cell.text)
        if tip_parts:
            pending_tip = " ".join(tip_parts).strip()

        measures_right_limit = (cantidad_x - 5) if cantidad_x is not None else (precio_x - 5)
        measures_parts: List[str] = []
        if medidas_x is not None:
            for cell in row.cells:
                if cell.x < (medidas_x - 20):
                    continue
                if cell.x >= measures_right_limit:
                    continue
                if re.search(r"\d+\s*x\s*\d+", cell.text or "", flags=re.IGNORECASE):
                    measures_parts.append(cell.text)
        if measures_parts:
            pending_measures = " ".join(measures_parts).strip()
        elif not pending_measures:
            measure_match = re.search(r"\b\d+\s*x\s*\d+\b", " ".join(cell.text for cell in row.cells), flags=re.IGNORECASE)
            if measure_match:
                pending_measures = measure_match.group(0)

        if cantidad_x is not None:
            qty_value: int | None = None
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 5):
                    continue
                nums = re.findall(r"\d+", cell.text or "")
                if not nums:
                    continue
                qty_value = int(nums[0])
                break
            if qty_value is not None:
                pending_qty = qty_value

        glass_text_parts: List[str] = []
        glass_left_limit = vidrio_x - 5
        glass_right_limit = (medidas_x - 20) if medidas_x is not None else (precio_x - 5)
        for cell in row.cells:
            if cell.x < glass_left_limit:
                continue
            if cell.x >= glass_right_limit:
                continue
            low = _normalize(cell.text)
            if low in header_tokens:
                continue
            if any(ch.isalpha() for ch in cell.text):
                glass_text_parts.append(cell.text)
        if glass_text_parts:
            pending_glass_parts.extend(glass_text_parts)

        price_value: float | None = None
        for cell in row.cells:
            if cell.x < (precio_x - 20):
                continue
            parsed = _parse_number(cell.text)
            if parsed is not None:
                price_value = parsed
                break

        if price_value is None and cantidad_x is not None:
            for cell in row.cells:
                if cell.x < (cantidad_x - 20):
                    continue
                if cell.x >= (precio_x - 20):
                    continue
                nums = re.findall(r"\d+(?:[.,]\d+)?", cell.text or "")
                if len(nums) < 2:
                    continue
                parsed_last = _parse_number(nums[-1])
                if parsed_last is not None:
                    price_value = parsed_last
                    break

        if price_value is None:
            continue

        glass_name = " ".join(part for part in pending_glass_parts if str(part).strip()).strip()
        if not glass_name:
            raw_candidates = []
            for cell in row.cells:
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalpha() for ch in cell.text):
                    raw_candidates.append(cell.text)
            glass_name = " ".join(raw_candidates).strip()
        if not glass_name:
            pending_tip = ""
            pending_measures = ""
            pending_qty = None
            pending_glass_parts = []
            continue

        if not pending_tip:
            for cell in row.cells:
                if cell.x >= (vidrio_x - 5):
                    continue
                low = _normalize(cell.text)
                if low in header_tokens:
                    continue
                if any(ch.isalnum() for ch in cell.text):
                    pending_tip = cell.text.strip()
                    break

        lado1, lado2 = _extract_measure_sides(pending_measures or glass_name)
        vidrio1_raw, vidrio2_raw, camara = _split_pv_glass_components(glass_name)
        if camara is None:
            pending_tip = ""
            pending_measures = ""
            pending_qty = None
            pending_glass_parts = []
            continue

        vidrio1 = _map_pv_glass_code(vidrio1_raw)
        vidrio2 = _map_pv_glass_code(vidrio2_raw)

        detail_rows.append(
            {
                "TIP.": pending_tip,
                "LADO1 VIDRIO": lado1,
                "LADO2 VIDRIO": lado2,
                "CANT.": pending_qty,
                "VIDRIO1": vidrio1,
                "VIDRIO2": vidrio2,
                "CAMARA": camara,
            }
        )

        pending_tip = ""
        pending_measures = ""
        pending_qty = None
        pending_glass_parts = []

    return detail_rows


def _find_row_by_label(rows: List[Row], label: str) -> Row:
    needle = _normalize(label).rstrip(":")
    for row in rows:
        for cell in row.cells:
            if _normalize(cell.text).startswith(needle):
                return row
    raise RuntimeError(f"No se encontro la fila '{label}' en INFORME.pdf")

def _extract_pa_totals(pa_pdf: Path) -> Dict[str, float]:
    """Extrae el Precio Neto y Total final del Pedido de Accesorios."""
    cells = _extract_cells(pa_pdf)
    rows = _group_rows(cells)
    
    precio = 0.0
    total = 0.0
    
    for row in rows:
        text = _normalize(" ".join(c.text for c in row.cells))
        vals = _row_numeric_values(row)
        
        if "precio" in text and vals:
            precio = vals[-1]
        elif "total" in text and vals:
            total = vals[-1]
            
    # Fallback de seguridad: A veces pdfminer separa las etiquetas de los 
    # montos en distintas alturas (coordenada Y) dejándolos en filas distintas.
    if total == 0.0:
        all_numbers = []
        for cell in sorted(cells, key=lambda c: (-c.y, c.x)):
            num = _parse_number(cell.text)
            if num is not None:
                all_numbers.append(num)
        
        # Filtramos números grandes ignorando paginación o fechas
        large_nums = [n for n in all_numbers if n > 1000]
        if len(large_nums) >= 3:
            # Los últimos números suelen ser Precio, Iva, Total
            total = large_nums[-1]
            precio = large_nums[-3]
            
    return {"precio": precio, "total": total}


def _extract_informe_summary(informe_pdf: Path) -> Dict[str, float | None]:
    rows = _group_rows(_extract_cells(informe_pdf))

    accesorios_row = _find_row_by_label(rows, "Accesorios en Pesos")
    vidrios_row = _find_row_by_label(rows, "Vidrios en Pesos")

    accesorios_vals = _row_numeric_values(accesorios_row)
    vidrios_vals = _row_numeric_values(vidrios_row)

    if len(accesorios_vals) < 1:
        raise RuntimeError("No se pudo leer Consumo de 'Accesorios en Pesos'")
    if len(vidrios_vals) < 1:
        raise RuntimeError("No se pudo leer Consumo de 'Vidrios en Pesos'")

    return {
        "kg_consumidos": None,
        "kg_comprados": None,
        "accesorios_resto": accesorios_vals[0],
        "vidrios_pesos_consumo": vidrios_vals[0],
    }


def _round_to_int(value: float) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _write_integer(ws, row: int, col: int, value: float) -> None:
    cell = ws.cell(row=row, column=col, value=_round_to_int(float(value)))
    cell.number_format = "0"


def _write_integer_or_blank(ws, row: int, col: int, value: float | None) -> None:
    if value is None:
        ws.cell(row=row, column=col, value="")
        return
    _write_integer(ws, row, col, value)


def _auto_width(ws, min_width: int = 12, max_width: int = 95) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _write_excel(
    output_path: Path,
    perfiles: List[Dict[str, object]],
    retazos_resto: List[Dict[str, object]],
    retazos_alta_prestacion: List[Dict[str, object]],
    pv_detail_rows: List[Dict[str, object]],
    incluir_retazos: bool,
    bipuntos_cantidad: int,
    pv_totals: Dict[str, float],
    informe_summary: Dict[str, float | None] | None,
    pa_totals: Dict[str, float] | None = None,
) -> Dict[str, float | str]:
    
    total_vidrios_pv = sum(pv_totals.get(k, 0.0) for k in GLASS_ORDER)
    total_vidrios_pv_int = _round_to_int(total_vidrios_pv)
    
    if informe_summary:
        total_vidrios_informe = float(informe_summary["vidrios_pesos_consumo"])
        total_vidrios_informe_int = _round_to_int(total_vidrios_informe)
        diferencia = total_vidrios_pv_int - total_vidrios_informe_int
        coincide = diferencia == 0
        estado = "OK" if coincide else f"NO COINCIDE (dif: {diferencia:+d})"
    else:
        total_vidrios_informe_int = 0
        diferencia = 0
        estado = "NO APLICA (Sin INFORME)"

    wb = Workbook()
    ws = wb.active
    ws.title = "Control"
    bold = Font(bold=True)

    row = 1
    ws.cell(row=row, column=1, value="Perfiles (Opti.htm)").font = bold
    row += 1
    for col, header in enumerate(["CODIGO", "COLOR", "CANTIDAD"], start=1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    for item in perfiles:
        ws.cell(row=row, column=1, value=int(item["CODIGO"]))
        ws.cell(row=row, column=2, value=str(item["COLOR"]))
        ws.cell(row=row, column=3, value=int(item["CANTIDAD"]))
        row += 1

    if incluir_retazos:
        row += 1
        ws.cell(row=row, column=1, value="Retazos Utiles - Resto (Opti.htm)").font = bold
        row += 1
        for col, header in enumerate(["CODIGO", "COLOR", "LARGO RETAZO", "CANTIDAD"], start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1

        for item in retazos_resto:
            ws.cell(row=row, column=1, value=int(item["CODIGO"]))
            ws.cell(row=row, column=2, value=str(item["COLOR"]))
            ws.cell(row=row, column=3, value=int(item["LARGO_RETAZO"]))
            ws.cell(row=row, column=4, value=int(item["CANTIDAD"]))
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Retazos Alta Prestacion (Opti.htm)").font = bold
        row += 1
        for col, header in enumerate(["CODIGO", "COLOR", "LARGO RETAZO", "CANTIDAD"], start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1

        for item in retazos_alta_prestacion:
            ws.cell(row=row, column=1, value=int(item["CODIGO"]))
            ws.cell(row=row, column=2, value=str(item["COLOR"]))
            ws.cell(row=row, column=3, value=int(item["LARGO_RETAZO"]))
            ws.cell(row=row, column=4, value=int(item["CANTIDAD"]))
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Detalle Vidrios (PV.pdf)").font = bold
    row += 1
    for col, header in enumerate(PV_DETAIL_HEADERS, start=1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    for item in pv_detail_rows:
        for col, header in enumerate(PV_DETAIL_HEADERS, start=1):
            value = item.get(header, "")
            ws.cell(row=row, column=col, value="" if value is None else value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Resumen Vidrios (PV.pdf)").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Categoria").font = bold
    ws.cell(row=row, column=2, value="Precio").font = bold
    row += 1

    for category in GLASS_ORDER:
        ws.cell(row=row, column=1, value=category)
        _write_integer(ws, row, 2, float(pv_totals.get(category, 0.0)))
        row += 1

    row += 1
    if informe_summary:
        ws.cell(row=row, column=1, value="Resumen Informe (INFORME.pdf)").font = bold
        row += 1
        ws.cell(row=row, column=1, value="Concepto").font = bold
        ws.cell(row=row, column=2, value="Valor").font = bold
        row += 1

        ws.cell(row=row, column=1, value="Accesorios resto")
        _write_integer(ws, row, 2, float(informe_summary["accesorios_resto"]))
        row += 1

        ws.cell(row=row, column=1, value="Bipuntos")
        _write_integer(ws, row, 2, float(bipuntos_cantidad * BIPUNTO_UNITARIO))
        row += 1

        ws.cell(row=row, column=1, value="Kg consumidos")
        _write_integer_or_blank(ws, row, 2, informe_summary["kg_consumidos"])
        row += 1

        ws.cell(row=row, column=1, value="Kg comprados")
        _write_integer_or_blank(ws, row, 2, informe_summary["kg_comprados"])
        row += 1

        ws.cell(row=row, column=1, value="Vidrios en Pesos (INFORME - Consumo)")
        _write_integer(ws, row, 2, total_vidrios_informe)
        row += 1

        ws.cell(row=row, column=1, value="Total Vidrios (PV)")
        _write_integer(ws, row, 2, total_vidrios_pv)
        row += 1

        ws.cell(row=row, column=1, value="Chequeo Vidrios (PV vs INFORME)")
        ws.cell(row=row, column=2, value=estado)
    elif pa_totals:
        ws.cell(row=row, column=1, value="Resumen Accesorios (PA.pdf)").font = bold
        row += 1
        ws.cell(row=row, column=1, value="Concepto").font = bold
        ws.cell(row=row, column=2, value="Valor").font = bold
        row += 1

        ws.cell(row=row, column=1, value="Accesorios (Precio Neto)")
        _write_integer(ws, row, 2, pa_totals["precio"])
        row += 1
        
        ws.cell(row=row, column=1, value="Accesorios (Total final)")
        _write_integer(ws, row, 2, pa_totals["total"])
        row += 1

        ws.cell(row=row, column=1, value="Bipuntos")
        _write_integer(ws, row, 2, float(bipuntos_cantidad * BIPUNTO_UNITARIO))
        row += 1

        ws.cell(row=row, column=1, value="Total Vidrios (PV)")
        _write_integer(ws, row, 2, total_vidrios_pv)
        row += 1

    checklist = wb.create_sheet("Checklist")
    checklist.cell(row=1, column=1, value="Pregunta").font = bold
    checklist.cell(row=1, column=2, value="Chequeado").font = bold

    check_validation = DataValidation(
        type="list",
        formula1=f'"{CHECK_EMPTY},{CHECK_OK},{CHECK_NO}"',
        allow_blank=False,
    )
    checklist.add_data_validation(check_validation)

    for idx, pregunta in enumerate(CHECKLIST_PREGUNTAS, start=2):
        checklist.cell(row=idx, column=1, value=pregunta)
        check_cell = checklist.cell(row=idx, column=2, value=CHECK_EMPTY)
        check_cell.alignment = Alignment(horizontal="center", vertical="center")
        check_validation.add(check_cell)

    _auto_width(ws)
    _auto_width(checklist, min_width=14, max_width=130)
    wb.save(output_path)

    return {
        "total_vidrios_pv": total_vidrios_pv_int,
        "total_vidrios_informe": total_vidrios_informe_int,
        "diferencia": diferencia,
        "estado": estado,
    }


def _prompt_new_or_existing() -> str:
    """Retorna 'N' (nuevo), 'E' (existente via dialogo) o 'U' (ultimo generado)."""
    while True:
        choice = input(
            "¿Nuevo Excel [N] / Abrir existente [E] / Usar el último generado [U]? "
        ).strip().upper()
        if choice in ("N", "NUEVO"):
            return "N"
        if choice in ("E", "EXISTENTE"):
            return "E"
        if choice in ("U", "ULTIMO", "ÚLTIMO"):
            return "U"
        print("Responder N, E o U.")


def _prompt_select_excel_dialog(initial_dir: Path) -> Path:
    """Abre el diálogo nativo de Windows para elegir un .xlsx."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        raise SystemExit("tkinter no disponible. Instalá Python con soporte para tkinter.")
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path_str = filedialog.askopenfilename(
        title="Seleccionar Excel de cotización",
        initialdir=str(initial_dir),
        filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
    )
    root.destroy()
    if not path_str:
        raise SystemExit("No se seleccionó ningún archivo.")
    return Path(path_str)


def _find_last_excel(search_dir: Path) -> Path:
    """Devuelve el .xlsx modificado más recientemente en search_dir (excluye templates y temps)."""
    EXCLUDED_NAMES = {"Cotizador V7.xlsx"}
    candidates = [
        f for f in search_dir.glob("*.xlsx")
        if not f.name.startswith("~$") and f.name not in EXCLUDED_NAMES
    ]
    if not candidates:
        raise SystemExit(f"No se encontraron archivos .xlsx en {search_dir}")
    return max(candidates, key=lambda f: f.stat().st_mtime)


def _prompt_output_name(default_dir: Path) -> Path:
    """Pide nombre para el nuevo Excel y devuelve el Path completo."""
    while True:
        name = input("Nombre para el nuevo Excel (sin extensión): ").strip()
        if name:
            return default_dir / f"{name}.xlsx"
        print("El nombre no puede estar vacío.")


def _load_workbook_with_retry(path: Path):
    """Carga el workbook; si está abierto en Excel, pide cerrarlo y reintenta."""
    while True:
        try:
            return load_workbook(path)
        except PermissionError:
            input(
                f"\nEl archivo '{path.name}' está abierto en Excel.\n"
                "Cerralo y presioná Enter para reintentar... "
            )


def _prompt_variant_name() -> str:
    """Solicita el nombre de la variante (no puede estar vacío)."""
    while True:
        name = input("Nombre de la variante (ej: 'DVH dorm', 'VS M45'): ").strip()
        if name:
            return name
        print("El nombre no puede estar vacío.")


def _find_latest_base(bases_dir: Path):
    """Devuelve el Base_*.xlsx más reciente en la carpeta, o None si no hay."""
    candidates = sorted(bases_dir.glob("Base_*.xlsx"), reverse=True)
    return candidates[0] if candidates else None


def _prompt_select_base(bases_dir: Path) -> Path:
    """Lista las bases disponibles y deja elegir. Si hay una sola, la usa directamente."""
    bases = sorted(bases_dir.glob("Base_*.xlsx"), reverse=True)
    if not bases:
        raise SystemExit(f"No hay archivos Base_*.xlsx en {bases_dir}")
    if len(bases) == 1:
        print(f"Base de precios: {bases[0].name}")
        return bases[0]
    print("\nBases de precios disponibles:")
    for i, b in enumerate(bases, 1):
        marker = " <- más reciente" if i == 1 else ""
        print(f"  {i}. {b.name}{marker}")
    while True:
        resp = input(f"Elegir base [1-{len(bases)}, Enter = más reciente]: ").strip()
        if resp == "":
            return bases[0]
        if resp.isdigit() and 1 <= int(resp) <= len(bases):
            return bases[int(resp) - 1]
        print(f"Ingresar un número entre 1 y {len(bases)}, o Enter.")


def _read_prices_from_base(base_path: Path) -> dict:
    """Lee precios c/IVA de Base_*.xlsx. Devuelve {codigo: {'A': float, 'B': float, 'N': float}}."""
    wb_base = load_workbook(base_path, data_only=True)
    ws = wb_base["Precios"]
    prices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_padded = tuple(row) + (None,) * 5
        codigo, _desc, a_iva, b_iva, n_iva = row_padded[:5]
        if codigo is not None:
            prices[str(codigo).strip()] = {"A": a_iva, "B": b_iva, "N": n_iva}
    return prices


def _write_prices_to_perfiles(ws_bp, prices: dict) -> None:
    """Escribe precios c/IVA en Base perfiles (cols 12=A, 13=B, 14=N) buscando por código (col A)."""
    for row in ws_bp.iter_rows(min_row=4, max_col=14):
        codigo = row[0].value
        if not codigo:
            continue
        p = prices.get(str(codigo).strip())
        if p:
            ws_bp.cell(row=row[0].row, column=12).value = p["A"]
            ws_bp.cell(row=row[0].row, column=13).value = p["B"]
            ws_bp.cell(row=row[0].row, column=14).value = p["N"]


def _fill_cotizador_variant(
    wb: "Workbook",
    variant: str,
    perfiles: List[Dict[str, object]],
    retazos_resto: List[Dict[str, object]],
    retazos_alta_prestacion: List[Dict[str, object]],
    dvh_rows: List[Dict[str, object]],
    incluir_retazos: bool,
    bipuntos_cantidad: int,
    pv_totals: Dict[str, float],
    color: str,
    pa_totals: Dict[str, float] | None = None,
    informe_summary: Dict[str, float | None] | None = None,
    base_name: str | None = None,
    templado_rows: List[Dict[str, object]] | None = None,
) -> None:
    """Agrega hojas de una variante al workbook copiando las plantillas del V7."""
    pedido_name = f"Pedido {variant}"[:31]
    retazo_name = f"Retazo {variant}"[:31]
    dvh_name = f"DVH {variant}"[:31]
    principal_name = f"Principal {variant}"[:31]

    for name in [pedido_name, principal_name]:
        if name in wb.sheetnames:
            raise SystemExit(f"Ya existe una hoja '{name}'. Elegir un nombre diferente para la variante.")
    if incluir_retazos and retazo_name in wb.sheetnames:
        raise SystemExit(f"Ya existe una hoja '{retazo_name}'.")
    if dvh_rows and dvh_name in wb.sheetnames:
        raise SystemExit(f"Ya existe una hoja '{dvh_name}'.")

    # --- Pedido ---
    ws_pedido = wb.copy_worksheet(wb["Plantilla Pedido"])
    ws_pedido.title = pedido_name
    # Reemplazar referencias de tabla con rangos fijos para evitar nombres duplicados
    ws_pedido["K2"] = "=SUM(F2:F46)"
    ws_pedido["L2"] = "=SUM(G2:G46)"
    ws_pedido["M2"] = "=SUM(H2:H46)"
    # Limpiar TODAS las filas (A, B, C) y escribir solo los perfiles del Opti
    for r in range(2, ws_pedido.max_row + 1):
        ws_pedido.cell(row=r, column=1).value = None
        ws_pedido.cell(row=r, column=2).value = None
        ws_pedido.cell(row=r, column=3).value = None
    for i, item in enumerate(sorted(perfiles, key=lambda x: int(x["CODIGO"])), start=2):
        ws_pedido.cell(row=i, column=1).value = int(item["CODIGO"])
        ws_pedido.cell(row=i, column=2).value = color
        ws_pedido.cell(row=i, column=3).value = int(item["CANTIDAD"])

    # --- Retazo (solo si se analizaron retazos) ---
    retazo_created = False
    if incluir_retazos:
        ws_retazo = wb.copy_worksheet(wb["Plantilla Retazo"])
        ws_retazo.title = retazo_name
        ws_retazo["J4"] = "=SUM(H4:H149)"
        ws_retazo["V4"] = "=SUM(T4:T179)"
        RETAZO_DATA_START = 4
        RETAZO_MAX = ws_retazo.max_row
        for r in range(RETAZO_DATA_START, RETAZO_MAX + 1):
            for c in (2, 3, 4, 5):
                ws_retazo.cell(row=r, column=c).value = None
        for i, item in enumerate(retazos_resto):
            r = RETAZO_DATA_START + i
            ws_retazo.cell(row=r, column=2).value = int(item["CODIGO"])
            ws_retazo.cell(row=r, column=3).value = str(item["COLOR"])
            ws_retazo.cell(row=r, column=4).value = int(item["LARGO_RETAZO"])
            ws_retazo.cell(row=r, column=5).value = int(item["CANTIDAD"])
        for r in range(RETAZO_DATA_START, RETAZO_MAX + 1):
            for c in (14, 15, 16, 17):
                ws_retazo.cell(row=r, column=c).value = None
        for i, item in enumerate(retazos_alta_prestacion):
            r = RETAZO_DATA_START + i
            ws_retazo.cell(row=r, column=14).value = int(item["CODIGO"])
            ws_retazo.cell(row=r, column=15).value = str(item["COLOR"])
            ws_retazo.cell(row=r, column=16).value = int(item["LARGO_RETAZO"])
            ws_retazo.cell(row=r, column=17).value = int(item["CANTIDAD"])
        retazo_created = True

    # --- DVH (solo si hay vidrios DVH en la obra) ---
    dvh_created = False
    if dvh_rows:
        ws_dvh = wb.copy_worksheet(wb["Plantilla DVH"])
        ws_dvh.title = dvh_name
        ws_dvh["W2"] = "=SUM(P3:P30)"
        ws_dvh["W3"] = "=SUM(T3:T30)"
        DVH_DATA_START = 3
        DVH_MAX = ws_dvh.max_row
        for r in range(DVH_DATA_START, DVH_MAX + 1):
            for c in range(1, 8):
                ws_dvh.cell(row=r, column=c).value = None
        dvh_cols = ["TIP.", "LADO1 VIDRIO", "LADO2 VIDRIO", "CANT.", "VIDRIO1", "VIDRIO2", "CAMARA"]
        for i, item in enumerate(dvh_rows):
            r = DVH_DATA_START + i
            if r > DVH_MAX:
                print(
                    f"Advertencia: plantilla DVH soporta {DVH_MAX - DVH_DATA_START + 1} filas max.",
                    file=sys.stderr,
                )
                break
            for col_idx, header in enumerate(dvh_cols, start=1):
                ws_dvh.cell(row=r, column=col_idx).value = item.get(header)
        dvh_created = True

    # --- Templados (solo si hay vidrios simples/cortados en la obra) ---
    templados_name = f"Templados {variant}"[:31]
    templados_created = False
    if templado_rows:
        ws_tpl_glass = wb.copy_worksheet(wb["Plantilla Templados"])
        ws_tpl_glass.title = templados_name
        ws_tpl_glass["W2"] = "=SUM(I3:I30)"
        TPLG_DATA_START = 3
        TPLG_MAX = ws_tpl_glass.max_row
        for r in range(TPLG_DATA_START, TPLG_MAX + 1):
            for c in range(1, 6):
                ws_tpl_glass.cell(row=r, column=c).value = None
        tplg_cols = ["TIP.", "LADO1 VIDRIO", "LADO2 VIDRIO", "CANT.", "VIDRIO1"]
        for i, item in enumerate(templado_rows):
            r = TPLG_DATA_START + i
            if r > TPLG_MAX:
                print(
                    f"Advertencia: plantilla Templados soporta {TPLG_MAX - TPLG_DATA_START + 1} filas max.",
                    file=sys.stderr,
                )
                break
            for col_idx, header in enumerate(tplg_cols, start=1):
                ws_tpl_glass.cell(row=r, column=col_idx).value = item.get(header)
        templados_created = True

    # --- Principal: copia y actualiza referencias de hoja ---
    ws_principal = wb.copy_worksheet(wb["Plantilla Principal"])
    ws_principal.title = principal_name
    sheet_map = {"Plantilla Pedido": pedido_name}
    if retazo_created:
        sheet_map["Plantilla Retazo"] = retazo_name
    if templados_created:
        sheet_map["Plantilla Templados"] = templados_name
    for row in ws_principal.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                new_val = cell.value
                for old, new in sheet_map.items():
                    new_val = new_val.replace(f"'{old}'!", f"'{new}'!")
                cell.value = new_val
    # Celdas de input
    if informe_summary:
        ws_principal["B3"] = _round_to_int(float(informe_summary["accesorios_resto"]))
    elif pa_totals:
        ws_principal["B3"] = _round_to_int(pa_totals["precio"])
    ws_principal["B4"] = bipuntos_cantidad * BIPUNTO_UNITARIO
    ws_principal["B6"] = _round_to_int(float(pv_totals.get(GLASS_LAMIN, 0.0)))
    ws_principal["B7"] = _round_to_int(float(pv_totals.get(GLASS_CRUDOS, 0.0)))
    if dvh_created:
        ws_principal["B8"] = f"='{dvh_name}'!W2"
    else:
        ws_principal["B8"] = 0
    if templados_created:
        ws_principal["B9"] = f"='{templados_name}'!W2"
    else:
        ws_principal["B9"] = 0
    if base_name:
        ws_principal["F1"] = f"Base precios: {base_name}"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Control de cotizaciones: perfiles (Opti), vidrios (PV), Accesorios (PA) o Informe."
    )
    parser.add_argument("--pv", type=Path, default=Path("PV.pdf"), help="Ruta a PV.pdf")
    parser.add_argument("--informe", type=Path, default=Path("INFORME.pdf"), help="Ruta a INFORME.pdf")
    parser.add_argument("--pa", type=Path, default=Path("PA.pdf"), help="Ruta a PA.pdf (Pedido de Accesorios)")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("Control_Cotizacion.xlsx"),
        help="Excel de salida",
    )
    parser.add_argument("--color", type=str, default=None, help="Color de obra: N/B/A o nombre completo")
    parser.add_argument("--alta", type=str, default=None, help="Compatibilidad hacia atras.")
    parser.add_argument("--retazos", type=str, default=None, help="Analizar retazos utiles desde Opti.htm? S/N")
    parser.add_argument("--opti", type=Path, default=Path(r"C:\MDT_Winproject_2\Proyecto\Opti.htm"), help="Ruta al archivo Opti.htm")
    parser.add_argument("--bipuntos", type=str, default=None, help="Cantidad de bipuntos (entero >= 0)")
    parser.add_argument("--check_vidrios", type=str, default=None, help="Chequear vidrios con INFORME? S/N")
    parser.add_argument(
        "--cotizador",
        type=Path,
        default=None,
        help="Ruta al Cotizador V7.xlsx (template). Si no se especifica, busca 'Cotizador V7.xlsx' en el directorio actual.",
    )
    return parser.parse_args()


def main() -> int:
    args = _parse_args()

    color = _prompt_color(args.color)
    bipuntos = _prompt_non_negative_int(
        args.bipuntos,
        (
            "Cuantos Bipuntos tiene la obra?\n"
            "Corredizas de 2 y 3 hojas: 2 bipuntos\n"
            "Corredizas de 4 y 6 hojas: 3 bipuntos\n"
            "Las corredizas de alta prestacion YA INCLUYEN y no debe contarlas nuevamente.\n"
            "Ingrese cantidad total de bipuntos: "
        ),
    )
    analizar_retazos = _prompt_yes_no(
        args.retazos if args.retazos is not None else args.alta,
        "Analizar retazos utiles desde Opti.htm? [S/N]: ",
    )

    # Validacion de archivos
    if not args.pv.exists():
        raise SystemExit(f"No existe el archivo: {args.pv}")
    if not args.opti.exists():
        raise SystemExit(f"No existe el archivo Opti.htm en: {args.opti}")
    if not args.pa.exists():
        raise SystemExit(f"No existe el archivo de Accesorios: {args.pa}")

    # Extracciones
    retazos_resto: List[Dict[str, object]] = []
    retazos_alta_prestacion: List[Dict[str, object]] = []
    if analizar_retazos:
        retazos_resto, retazos_alta_prestacion = _extract_split_retazos(args.opti, color)

    perfiles = _extract_opti_profiles(args.opti, color)
    if not perfiles:
        print("Aviso: no se encontraron perfiles en Opti.htm", file=sys.stderr)

    pv_detail_rows = _extract_pv_detailed_rows(args.pv)
    dvh_rows = [r for r in pv_detail_rows if r.get("CAMARA") not in (None, "")]
    templado_rows = [r for r in pv_detail_rows if r.get("CAMARA") in (None, "")]
    pv_totals = _extract_pv_summary(args.pv)

    informe_summary = None
    pa_totals = _extract_pa_totals(args.pa)

    # Detectar template del Cotizador V7
    cotizador_template = args.cotizador
    if cotizador_template is None:
        default_v7 = Path("Cotizador V7.xlsx")
        if default_v7.exists():
            cotizador_template = default_v7

    if cotizador_template is not None:
        if not cotizador_template.exists():
            raise SystemExit(f"No existe el archivo Cotizador: {cotizador_template}")

        # Preguntar nuevo Excel / existente / último
        opcion = _prompt_new_or_existing()
        variant = _prompt_variant_name()

        output_dir = args.out.parent.resolve() if str(args.out.parent) != "." else Path.cwd()

        if opcion == "N":
            output_path = _prompt_output_name(output_dir)
            while True:
                try:
                    shutil.copy(cotizador_template, output_path)
                    break
                except PermissionError:
                    input(
                        f"\nEl archivo '{output_path.name}' está abierto en Excel.\n"
                        "Cerralo y presioná Enter para reintentar... "
                    )
            wb = _load_workbook_with_retry(output_path)
        elif opcion == "E":
            output_path = _prompt_select_excel_dialog(output_dir)
            wb = _load_workbook_with_retry(output_path)
        else:  # "U"
            output_path = _find_last_excel(output_dir)
            print(f"Usando último Excel generado: {output_path.name}")
            wb = _load_workbook_with_retry(output_path)

        # Si el archivo no tiene las hojas plantilla (ej: creado con versión anterior),
        # las importa automáticamente desde el template V7.
        if "Plantilla Pedido" not in wb.sheetnames:
            print(f"El archivo no tiene hojas plantilla — importando desde {cotizador_template.name}...")
            wb_tpl = load_workbook(cotizador_template)
            for tpl_name in [
                "Plantilla Principal", "Base perfiles", "Base vidrios",
                "Plantilla Retazo", "Plantilla Pedido", "Plantilla DVH", "Plantilla Templados",
            ]:
                if tpl_name not in wb.sheetnames and tpl_name in wb_tpl.sheetnames:
                    src = wb_tpl[tpl_name]
                    dst = wb.create_sheet(tpl_name)
                    for row in src.iter_rows():
                        for cell in row:
                            dst.cell(row=cell.row, column=cell.column).value = cell.value
            print("  Hojas plantilla importadas correctamente.")

        # Cargar precios desde base central si existe
        bases_dir = Path(__file__).parent / "Bases de Precios"
        base_name = None
        if bases_dir.exists():
            base_path = _prompt_select_base(bases_dir)
            prices = _read_prices_from_base(base_path)
            _write_prices_to_perfiles(wb["Base perfiles"], prices)
            base_name = base_path.stem
            print(f"Precios actualizados desde: {base_path.name}")
        else:
            print("Aviso: carpeta 'Bases de Precios/' no encontrada, usando precios del V7.")

        _fill_cotizador_variant(
            wb,
            variant,
            perfiles,
            retazos_resto,
            retazos_alta_prestacion,
            dvh_rows,
            analizar_retazos,
            bipuntos,
            pv_totals,
            color,
            pa_totals=pa_totals,
            informe_summary=informe_summary,
            base_name=base_name,
            templado_rows=templado_rows,
        )
        wb.save(output_path)
        print(f"Cotizador guardado: {output_path.resolve()}")
    else:
        _write_excel(
            args.out,
            perfiles,
            retazos_resto,
            retazos_alta_prestacion,
            pv_detail_rows,
            analizar_retazos,
            bipuntos,
            pv_totals,
            informe_summary=None,
            pa_totals=pa_totals,
        )
        print(f"Excel generado: {args.out.resolve()}")

    print(f"Color aplicado a perfiles: {color}")
    print(f"Bipuntos: {bipuntos} (costo total: {bipuntos * BIPUNTO_UNITARIO})")
    if pa_totals:
        print(f"Total Accesorios (Precio Neto): {_round_to_int(pa_totals['precio'])}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
